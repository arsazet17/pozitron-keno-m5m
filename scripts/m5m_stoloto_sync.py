#!/usr/bin/env python3
import asyncio, copy, json, os, re, sys
from datetime import datetime, date, time as dt_time, timezone, timedelta
from pathlib import Path
from openpyxl import load_workbook
from playwright.async_api import async_playwright

LOGIN_URL='https://oauth.stoloto.ru/login'
ARCHIVE_URL='https://m.stoloto.ru/keno2/archive/'
ARCHIVE_JSON=Path('data/archive.json')
ARCHIVE_XLSX=Path('data/m5m_stolby_po_date_vremeni.xlsx')
LAST_SYNC=Path('data/last_sync.json')
TAIL_SIZE=10
PAGE_READ_ATTEMPTS=3
SCHEDULE=[
'00:02','00:17','00:32','01:02','01:17','01:32','02:02','02:17','02:32','03:02','03:32',
'04:02','04:17','04:32','05:02','05:17','05:32','06:02','06:17','06:32','07:02','07:32',
'08:02','08:17','08:32','09:02','09:17','09:32','10:02','10:17','10:32','11:02','11:32',
'12:02','12:17','12:32','13:02','13:17','13:32','14:02','14:17','14:32','15:02','15:32',
'16:02','16:17','16:32','17:02','17:17','17:32','18:02','18:17','18:32','19:02','19:32',
'20:02','20:17','20:32','21:02','21:17','21:32','22:02','22:17','22:32','23:02','23:32']
SCHEDULE_SET=set(SCHEDULE)
MONTHS={'января':1,'февраля':2,'марта':3,'апреля':4,'мая':5,'июня':6,'июля':7,'августа':8,'сентября':9,'октября':10,'ноября':11,'декабря':12}

def norm(s): return re.sub(r'[ \t]+',' ',str(s or '').replace('\xa0',' ')).strip()
def moscow_today(): return (datetime.now(timezone.utc)+timedelta(hours=3)).date()

def parse_date_label(label):
    raw=norm(label).lower(); today=moscow_today()
    if raw=='сегодня': d=today
    elif raw=='вчера': d=today-timedelta(days=1)
    else:
        m=re.fullmatch(r'(\d{1,2})[./-](\d{1,2})[./-](\d{2,4})',raw)
        if m:
            y=int(m.group(3)); y=y+2000 if y<100 else y
            d=date(y,int(m.group(2)),int(m.group(1)))
        else:
            m=re.fullmatch(r'(\d{1,2})\s+([а-яё]+)(?:\s+(\d{4}))?',raw)
            if not m or m.group(2) not in MONTHS: return None
            y=int(m.group(3)) if m.group(3) else today.year; mm=MONTHS[m.group(2)]
            if not m.group(3) and mm>today.month+6: y-=1
            d=date(y,mm,int(m.group(1)))
    return d.strftime('%d.%m.%y')

def parse_archive_date(s):
    for fmt in ('%d.%m.%y','%d.%m.%Y'):
        try:return datetime.strptime(norm(s),fmt).date()
        except Exception:pass
    return None

def parse_time(text):
    m=re.search(r'\b([01]?\d|2[0-3]):([0-5]\d)(?::[0-5]\d)?\b',str(text or ''))
    return f'{int(m.group(1)):02d}:{m.group(2)}' if m else None

def parse_draw(text):
    m=re.search(r'№\s*([0-9]{4,})',str(text or ''))
    return int(m.group(1)) if m else None

def parse_column(text):
    m=re.search(r'столб(?:ец)?\s*[:№#-]?\s*([1-9]|10)\b',norm(text),re.I)
    return int(m.group(1)) if m else None

def valid_col(v):
    try:n=int(v)
    except Exception:return None
    return n if 1<=n<=10 else None

async def login(page,email,password):
    await page.goto(LOGIN_URL,wait_until='domcontentloaded',timeout=60000)
    login_loc=None; pass_loc=None
    for sel in ['input[type="email"]','input[name*="email" i]','input[name*="login" i]',
                'input[autocomplete="username"]','input[type="text"]']:
        loc=page.locator(sel).first
        if await loc.count(): login_loc=loc; break
    for sel in ['input[type="password"]','input[name*="password" i]',
                'input[autocomplete="current-password"]']:
        loc=page.locator(sel).first
        if await loc.count(): pass_loc=loc; break
    if login_loc is None or pass_loc is None:
        raise RuntimeError(f'OAuth fields not found; url={page.url}')
    await login_loc.fill(email); await pass_loc.fill(password)
    clicked=False
    for btn in [page.get_by_role('button',name=re.compile('войти',re.I)).first,
                page.locator('button[type="submit"]').first,
                page.locator('input[type="submit"]').first]:
        if await btn.count():
            await btn.click(); clicked=True; break
    if not clicked: raise RuntimeError('OAuth submit button not found')
    try: await page.wait_for_load_state('domcontentloaded',timeout=20000)
    except Exception: pass
    await page.wait_for_timeout(3500)

async def primary_dom_collect(page):
    raw=await page.locator('body').evaluate(r'''() => {
      const drawRx=/№\s*\d{4,}/;
      const dateRx=/^(Сегодня|Вчера|\d{1,2}[.\/-]\d{1,2}[.\/-]\d{2,4}|\d{1,2}\s+(?:января|февраля|марта|апреля|мая|июня|июля|августа|сентября|октября|ноября|декабря)(?:\s+\d{4})?)$/i;
      const norm=s=>String(s||'').replace(/\u00a0/g,' ').replace(/[ \t]+/g,' ').trim();
      const all=[...document.querySelectorAll('body *')];
      function nearestDate(el){
        let best=null;
        for(const node of all){
          if(node===el||el.contains(node))continue;
          const pos=node.compareDocumentPosition(el);
          if(!(pos&Node.DOCUMENT_POSITION_FOLLOWING))continue;
          const t=norm(node.innerText||node.textContent||'');
          if(!t||t.length>40||!dateRx.test(t))continue;
          if(node.children&&node.children.length>3)continue;
          best=t;
        }
        return best;
      }
      let rows=[...document.querySelectorAll('tr')].filter(el=>drawRx.test(el.innerText||''));
      if(!rows.length)rows=all.filter(el=>{
        const t=norm(el.innerText||'');
        return drawRx.test(t)&&![...el.children].some(ch=>drawRx.test(norm(ch.innerText||'')));
      });
      return rows.map(el=>({text:el.innerText||'',dateLabel:nearestDate(el)}));
    }''')
    out=[]; carry=None
    for row in raw:
        text=str(row.get('text','')); label=norm(row.get('dateLabel',''))
        if label: carry=label
        draw=parse_draw(text); tm=parse_time(text); col=parse_column(text)
        ds=parse_date_label(label or carry) if (label or carry) else None
        if draw and tm in SCHEDULE_SET and col and ds:
            out.append({'draw':draw,'date':ds,'time':tm,'column':col})
    uniq={x['draw']:x for x in out}
    return sorted(uniq.values(),key=lambda x:x['draw'])

def fallback_text_collect(body_text):
    lines=[norm(x) for x in str(body_text or '').splitlines()]
    lines=[x for x in lines if x]
    date_rx=re.compile(
        r'^(Сегодня|Вчера|\d{1,2}[./-]\d{1,2}[./-]\d{2,4}|'
        r'\d{1,2}\s+(?:января|февраля|марта|апреля|мая|июня|июля|августа|'
        r'сентября|октября|ноября|декабря)(?:\s+\d{4})?)$',re.I)
    out=[]; current_date=None
    for i,line in enumerate(lines):
        if date_rx.fullmatch(line):
            current_date=parse_date_label(line) or current_date
        if not re.search(r'№\s*\d{4,}',line): continue
        if current_date is None:
            for j in range(max(0,i-5),i):
                if date_rx.fullmatch(lines[j]):
                    current_date=parse_date_label(lines[j]) or current_date
        chunk=' '.join(lines[i:min(len(lines),i+8)])
        draw=parse_draw(chunk); tm=parse_time(chunk); col=parse_column(chunk)
        if draw and tm in SCHEDULE_SET and col and current_date:
            out.append({'draw':draw,'date':current_date,'time':tm,'column':col})
    uniq={x['draw']:x for x in out}
    return sorted(uniq.values(),key=lambda x:x['draw'])

async def collect(page):
    last_diag=None
    for page_attempt in range(1,PAGE_READ_ATTEMPTS+1):
        try:
            await page.goto(ARCHIVE_URL,wait_until='domcontentloaded',timeout=60000)
        except Exception as e:
            print(f'WARN: archive goto attempt {page_attempt}: {e}',file=sys.stderr)

        try:
            await page.wait_for_load_state('networkidle',timeout=12000)
        except Exception:
            pass
        await page.wait_for_timeout(2500 + 1000*page_attempt)

        primary=await primary_dom_collect(page)
        try:
            body=await page.locator('body').inner_text(timeout=10000)
        except Exception:
            body=''
        fallback=fallback_text_collect(body)

        merged={x['draw']:x for x in primary}
        for x in fallback:
            merged.setdefault(x['draw'],x)
        out=sorted(merged.values(),key=lambda x:x['draw'])

        try:title=await page.title()
        except Exception:title=''

        last_diag={
            'pageAttempt':page_attempt,
            'url':page.url,
            'title':title,
            'primary':len(primary),
            'fallback':len(fallback),
            'merged':len(out),
            'bodyHead':norm(body)[:700]
        }

        print(
            f"Stoloto page attempt {page_attempt}/{PAGE_READ_ATTEMPTS}: "
            f"url={page.url} primary={len(primary)} "
            f"fallback={len(fallback)} merged={len(out)}"
        )

        if len(out)>=TAIL_SIZE:
            return out[-TAIL_SIZE:]

        try:
            await page.reload(wait_until='domcontentloaded',timeout=60000)
        except Exception:
            pass
        await page.wait_for_timeout(1800)

    raise RuntimeError(
        f"Only {(last_diag or {}).get('merged',0)} recent draws found after "
        f"{PAGE_READ_ATTEMPTS} page reads; diagnostics="
        + json.dumps(last_diag,ensure_ascii=False)
    )

async def stable_tail(page):
    reads=[]
    for check in range(1,4):
        x=await collect(page)
        if len(x)<TAIL_SIZE:
            raise RuntimeError(f'Only {len(x)} recent draws found')
        reads.append({r['draw']:r for r in x})
        if check<3:
            await page.wait_for_timeout(900)

    common=sorted(set(reads[0])&set(reads[1])&set(reads[2]))[-TAIL_SIZE:]
    if len(common)<TAIL_SIZE:
        raise RuntimeError('Tail changed between checks')

    stable=[]
    for draw in common:
        a,b,c=reads[0][draw],reads[1][draw],reads[2][draw]
        if (a['date'],a['time'],a['column'])==(b['date'],b['time'],b['column'])==(c['date'],c['time'],c['column']):
            stable.append(a)

    if len(stable)<TAIL_SIZE:
        raise RuntimeError('Triple check failed')
    return stable

def header_map(rows):
    return {str(v):i for i,v in enumerate(rows[0]) if i>0 and v is not None}

def ensure_date_row(rows,ds):
    for i,r in enumerate(rows[1:],1):
        if str(r[0])==ds:return i
    row=[None]*len(rows[0]); row[0]=ds; rows.append(row)
    return len(rows)-1

def update_excel(added):
    if not added:return
    wb=load_workbook(ARCHIVE_XLSX)
    ws=wb[wb.sheetnames[0]]

    xh={}
    for c in range(2,ws.max_column+1):
        v=ws.cell(1,c).value
        if isinstance(v,(dt_time,datetime)):
            t=v.strftime('%H:%M')
        else:
            m=re.match(r'^(\d{1,2}):(\d{2})',norm(v))
            t=f'{int(m.group(1)):02d}:{m.group(2)}' if m else None
        if t:xh[t]=c

    rows_by={}
    for r in range(2,ws.max_row+1):
        v=ws.cell(r,1).value
        if isinstance(v,(date,datetime)):
            k=v.strftime('%d.%m.%y')
        else:
            d=parse_archive_date(v)
            k=d.strftime('%d.%m.%y') if d else norm(v)
        if k:rows_by[k]=r

    for x in added:
        c=xh.get(x['time'])
        if c is None:continue
        r=rows_by.get(x['date'])
        if r is None:
            r=ws.max_row+1; src=max(2,r-1)
            for cc in range(1,ws.max_column+1):
                ws.cell(r,cc)._style=copy.copy(ws.cell(src,cc)._style)
            ws.cell(r,1).value=x['date']; rows_by[x['date']]=r
        cur=valid_col(ws.cell(r,c).value)
        if cur is not None and cur!=x['column']:
            raise RuntimeError(f'Excel conflict {x}')
        ws.cell(r,c).value=x['column']

    wb.save(ARCHIVE_XLSX)

async def main():
    email=os.getenv('STOLOTO_EMAIL','').strip()
    password=os.getenv('STOLOTO_PASSWORD','').strip()
    if not email or not password:
        raise RuntimeError('Set STOLOTO_EMAIL and STOLOTO_PASSWORD secrets')

    async with async_playwright() as p:
        browser=await p.chromium.launch(headless=True)
        try:
            ctx=await browser.new_context(
                locale='ru-RU',
                timezone_id='Europe/Moscow',
                viewport={'width':390,'height':844}
            )
            page=await ctx.new_page()
            await login(page,email,password)
            stable=await stable_tail(page)
        finally:
            await browser.close()

    archive=json.loads(ARCHIVE_JSON.read_text(encoding='utf-8'))
    rows=archive['rows']; hm=header_map(rows); added=[]; confirmed=0

    for x in stable:
        c=hm.get(x['time'])
        if c is None:continue
        r=ensure_date_row(rows,x['date'])
        cur=valid_col(rows[r][c])

        if cur is None:
            rows[r][c]=x['column']; added.append(x)
        elif cur==x['column']:
            confirmed+=1
        else:
            raise RuntimeError(
                f"Archive conflict {x['date']} {x['time']}: {cur} != {x['column']}"
            )

    if added:
        ARCHIVE_JSON.write_text(
            json.dumps(archive,ensure_ascii=False,separators=(',',':'))+'\n',
            encoding='utf-8'
        )
        update_excel(added)

    last=stable[-1]
    LAST_SYNC.write_text(
        json.dumps({
            'updatedAt':datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace('+00:00','Z'),
            'source':'Stoloto OAuth M5M tail10 triple-check + fallback parser',
            'stableDraws':len(stable),
            'confirmedExisting':confirmed,
            'added':len(added),
            'latestOfficial':last,
            'addedRows':added
        },ensure_ascii=False,indent=2)+'\n',
        encoding='utf-8'
    )
    print(f"M5M PASS: added={len(added)}, confirmed={confirmed}, latest={last}")

if __name__=='__main__':
    try:
        asyncio.run(main())
    except Exception as e:
        print(f'FAIL: {e}',file=sys.stderr)
        raise
